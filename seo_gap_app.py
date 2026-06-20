"""
seo_gap_app.py  —  Streamlit arayüzü
Çalıştırma: streamlit run seo_gap_app.py
"""

import os, re, time
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from urllib.parse import urlparse
import xml.etree.ElementTree as ET

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import streamlit as st
from dotenv import load_dotenv

load_dotenv()

st.set_page_config(page_title="SEO Gap Analyzer", page_icon="🔍", layout="wide")

# ══════════════════════════════════════════════════════════════
# TASARIM — açık tema, veri-yoğun dashboard
# ══════════════════════════════════════════════════════════════

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #f7f8fa; color: #1a1d27; }
#MainMenu, header, footer { visibility: hidden; }
.block-container { padding-top: 2rem; max-width: 1280px; }

.app-header {
    border-bottom: 1px solid #e3e5ea; padding-bottom: 16px; margin-bottom: 24px;
}
.app-header h1 { font-size: 1.4rem; font-weight: 700; color: #14161f; margin: 0; letter-spacing: -0.01em; }
.app-header p { color: #767a8a; font-size: 0.85rem; margin: 2px 0 0; }

div[data-testid="stHorizontalBlock"] { align-items: stretch !important; }
.stTextInput input {
    background: #fff !important; border: 1px solid #d8dae2 !important; border-radius: 8px !important;
    color: #14161f !important; font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.92rem !important; height: 44px !important; padding: 0 14px !important;
}
.stTextInput input:focus { border-color: #4f6ef7 !important; box-shadow: 0 0 0 3px rgba(79,110,247,0.12) !important; }
.stButton button {
    background: #4f6ef7 !important; color: #fff !important; border: none !important;
    border-radius: 8px !important; height: 44px !important; font-weight: 600 !important;
    font-size: 0.9rem !important; width: 100% !important; margin-top: 0 !important;
}
.stButton button:hover { background: #3d5be0 !important; }

.metric-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; margin: 24px 0; }
.metric-card { background: #fff; border: 1px solid #e3e5ea; border-radius: 10px; padding: 16px 18px; }
.metric-card .label { font-size: 0.68rem; font-weight: 600; letter-spacing: 0.06em; text-transform: uppercase; color: #9296a3; margin-bottom: 6px; }
.metric-card .value { font-size: 1.65rem; font-weight: 700; color: #14161f; font-family: 'JetBrains Mono', monospace; line-height: 1; }
.metric-card .sub { font-size: 0.74rem; color: #9296a3; margin-top: 4px; }
.metric-card.warn .value { color: #d23c3c; }
.metric-card.good .value { color: #1a9b5c; }

.section-title { font-size: 0.78rem; font-weight: 700; letter-spacing: 0.04em; text-transform: uppercase; color: #4a4e5c; margin: 28px 0 12px; display: flex; align-items: center; gap: 8px; }
.section-title .count { background: #eceef3; color: #767a8a; font-size: 0.7rem; padding: 1px 7px; border-radius: 10px; font-weight: 600; }

.legend { display: flex; gap: 18px; margin-bottom: 12px; font-size: 0.78rem; color: #767a8a; }
.legend span.dot { display: inline-block; width: 9px; height: 9px; border-radius: 2px; margin-right: 6px; }

.dtable { background: #fff; border: 1px solid #e3e5ea; border-radius: 10px; overflow: hidden; }
.dtable .row { display: grid; padding: 11px 16px; border-bottom: 1px solid #eef0f4; align-items: center; font-size: 0.86rem; }
.dtable .head { background: #fafbfc; font-size: 0.68rem; font-weight: 700; letter-spacing: 0.05em; text-transform: uppercase; color: #9296a3; }
.cat-name { font-weight: 600; color: #14161f; }
.mono { font-family: 'JetBrains Mono', monospace; color: #4a4e5c; font-size: 0.84rem; }
.sample-kw { font-size: 0.74rem; color: #9296a3; padding: 0 16px 11px; border-bottom: 1px solid #eef0f4; }

.badge { display: inline-block; padding: 2px 9px; border-radius: 12px; font-size: 0.68rem; font-weight: 700; letter-spacing: 0.02em; }
.badge-opportunity { background: #e3f5ec; color: #1a9b5c; }
.badge-high  { background: #fdf0e0; color: #b8650f; }
.badge-none  { background: #fdeaea; color: #d23c3c; }

div[data-testid="stDownloadButton"] button {
    background: #fff !important; border: 1px solid #d8dae2 !important; color: #4a4e5c !important;
    border-radius: 8px !important; font-weight: 600 !important;
}

.warn-box { background: #fdf0e0; border: 1px solid #f0d4a8; border-radius: 8px; padding: 12px 16px; margin-bottom: 18px; font-size: 0.82rem; color: #8a5a10; }
.stProgress > div > div { background: #4f6ef7 !important; }
.streamlit-expanderHeader { font-size: 0.85rem !important; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# CREDENTIALS
# ══════════════════════════════════════════════════════════════

DFS_LOGIN    = os.getenv("DATAFORSEO_LOGIN", "")
DFS_PASSWORD = os.getenv("DATAFORSEO_PASSWORD", "")

if not DFS_LOGIN or not DFS_PASSWORD:
    st.error("⚠️ `.env` dosyasında DATAFORSEO_LOGIN ve DATAFORSEO_PASSWORD eksik.")
    st.stop()

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; SEOGapAnalyzer/1.0)"}


# ══════════════════════════════════════════════════════════════
# 1. SHOPIFY — kategoriler + TÜM ürün başlıkları
# ══════════════════════════════════════════════════════════════

def detect_platform(domain, progress_cb=None):
    for attempt in range(4):
        try:
            r = requests.get(f"https://{domain}/collections.json?limit=1", headers=HEADERS, timeout=10)
            if r.status_code == 200:
                try:
                    if "collections" in r.json():
                        return "shopify"
                except Exception:
                    pass
                return "generic"
            if r.status_code in (429, 503):
                wait = 5 * (attempt + 1)   # 5s, 10s, 15s, 20s — kademeli artan bekleme
                if progress_cb:
                    progress_cb(f"⚠️ Rate-limit (HTTP {r.status_code}) — {wait}s bekleniyor, deneme {attempt+1}/4...")
                time.sleep(wait)
                continue
            if r.status_code == 404:
                if progress_cb:
                    progress_cb("Shopify API bulunamadı (404) — site Shopify değil, genel tarama kullanılacak.")
                return "generic"
            if progress_cb:
                progress_cb(f"⚠️ Platform tespiti HTTP {r.status_code} döndü.")
            return "generic"
        except Exception as e:
            if progress_cb:
                progress_cb(f"⚠️ Platform tespit hatası: {e}")
            time.sleep(2)
    if progress_cb:
        progress_cb(
            "⚠️ Rate-limit nedeniyle platform tespit edilemedi. "
            "Bu genelde geçicidir — birkaç dakika bekleyip tekrar deneyin."
        )
    return "generic"


def fetch_all_collections(domain, progress_cb=None):
    collections, page = [], 1
    consecutive_errors = 0
    while True:
        cols = None
        for attempt in range(3):
            try:
                r = requests.get(f"https://{domain}/collections.json?limit=250&page={page}", headers=HEADERS, timeout=15)
                if r.status_code == 429:
                    wait = 4 * (attempt + 1)
                    if progress_cb:
                        progress_cb(f"⚠️ collections.json rate-limit (429) — {wait}s bekleniyor...")
                    time.sleep(wait)
                    continue
                if r.status_code != 200:
                    if progress_cb:
                        progress_cb(f"⚠️ collections.json HTTP {r.status_code} döndü (sayfa {page}).")
                    cols = []
                    break
                data = r.json()
                cols = data.get("collections", [])
                break
            except Exception as e:
                if progress_cb:
                    progress_cb(f"⚠️ collections.json hatası: {e}")
                cols = []
                break
        if cols is None:
            cols = []
        if not cols:
            consecutive_errors += 1
            if consecutive_errors >= 2 and page == 1:
                if progress_cb:
                    progress_cb("⚠️ İlk sayfada hiç koleksiyon bulunamadı — kategori verisi boş kalabilir.")
            break
        collections.extend(cols)
        if progress_cb:
            progress_cb(f"Koleksiyonlar çekiliyor... ({len(collections)} bulundu)")
        if len(cols) < 250:
            break
        page += 1
        time.sleep(0.3)
    return collections


def fetch_all_products(domain, progress_cb=None, max_pages=200):
    products, page = [], 1
    consecutive_errors = 0
    while page <= max_pages:
        retry_count = 0
        items = None
        while retry_count < 4:
            try:
                r = requests.get(f"https://{domain}/products.json?limit=250&page={page}", headers=HEADERS, timeout=20)
                if r.status_code == 429:
                    wait = 3 * (retry_count + 1)
                    if progress_cb and retry_count == 0:
                        progress_cb(f"⚠️ Rate limit (429) — {wait}s bekleniyor, tekrar denenecek...")
                    time.sleep(wait)
                    retry_count += 1
                    continue
                if r.status_code == 400:
                    # Shopify, sayfa limitinin sonuna gelindiğinde 400 döndürür — bu normal bir bitiş sinyali
                    items = []
                    break
                if r.status_code != 200:
                    consecutive_errors += 1
                    if progress_cb and consecutive_errors == 1:
                        progress_cb(f"⚠️ /products.json HTTP {r.status_code} döndü (sayfa {page}).")
                    items = []
                    break
                items = r.json().get("products", [])
                consecutive_errors = 0
                break
            except Exception as e:
                consecutive_errors += 1
                if progress_cb and consecutive_errors == 1:
                    progress_cb(f"⚠️ Ürün çekme hatası: {e}")
                items = []
                break
        if items is None:
            items = []
        if consecutive_errors >= 5:
            if progress_cb:
                progress_cb("⚠️ Çok fazla ardışık hata, ürün çekme durduruldu.")
            break
        if not items:
            break
        for p in items:
            products.append({
                "title": p.get("title", ""),
                "handle": p.get("handle", ""),
            })
        if progress_cb and page % 3 == 0:
            progress_cb(f"Ürünler çekiliyor... ({len(products)} ürün, sayfa {page})")
        if len(items) < 250:
            break
        page += 1
        time.sleep(1.2)   # Rate limit'e çarpmamak için sayfalar arası bekleme
    if progress_cb:
        progress_cb(f"Toplam {len(products)} ürün çekildi.")
    return products


NOISE_PATTERNS = [
    r'^\d+\s*(tl|%)', r'\btest\b', r'\bindirim\w*\b', r'\bkampanya\b',
    r'^\d+[a-z]?$', r'\boutlet\b', r'\bdiscount\b',
    r'^\d+\s*tl\s*(alt[ıi]|üzeri|ve)', r'^#',
    # Mağaza / lokasyon (AVM, outlet, şube, plaza) sayfaları — ürün kategorisi değil, marka bağımsız genel kural
    r'\bavm\b', r'\bmagaza\w*\b', r'\bmağaza\w*\b', r'\bpower outlet\b',
    r'\boutlet center\b', r'\bmarina\b', r'\bşube\w*\b', r'\bsube\w*\b',
    r'\bplaza\b', r'\bpark\b$', r'\bcity\b',
]
NOISE_RE = re.compile("|".join(NOISE_PATTERNS), re.IGNORECASE)


def build_categories(collections, progress_cb=None):
    real, noise = [], []
    for c in collections:
        title, handle = c.get("title", ""), c.get("handle", "")
        is_noise = bool(NOISE_RE.search(title) or NOISE_RE.search(handle))
        bucket = noise if is_noise else real
        bucket.append({
            "title": title,
            "handle": handle,
            "product_count": c.get("product_count"),   # generic scrape'ten geldiyse korunur, Shopify'da sonradan doldurulur
            "url": c.get("url"),
            "source": c.get("source", "shopify_api"),
        })
    if progress_cb:
        progress_cb(f"{len(real)} gerçek kategori, {len(noise)} kampanya/test/mağaza koleksiyonu ayrıldı.")
    return real, noise


def fetch_category_product_counts(domain, categories, progress_cb=None):
    """Her gerçek kategori için TAM ürün sayısını sayfalayarak çeker (250 üst sınırı yok)."""
    errors = 0
    for i, cat in enumerate(categories):
        total, page = 0, 1
        try:
            while True:
                retry_count = 0
                r = None
                while retry_count < 3:
                    r = requests.get(
                        f"https://{domain}/collections/{cat['handle']}/products.json?limit=250&page={page}",
                        headers=HEADERS, timeout=15
                    )
                    if r.status_code == 429:
                        time.sleep(2 * (retry_count + 1))
                        retry_count += 1
                        continue
                    break
                items = r.json().get("products", []) if r and r.status_code == 200 else []
                total += len(items)
                if len(items) < 250:
                    break
                page += 1
                time.sleep(0.4)
                if page > 20:
                    break
            cat["product_count"] = total
        except Exception:
            cat["product_count"] = 0
            errors += 1
        if progress_cb and i % 15 == 0:
            progress_cb(f"Kategori ürün sayıları hesaplanıyor... ({i+1}/{len(categories)})")
        time.sleep(0.25)
    if progress_cb and errors:
        progress_cb(f"⚠️ {errors} kategori için ürün sayısı çekilemedi.")
    return categories


# ── GENEL PLATFORM (Shopify olmayan siteler) ──────────────────

CATEGORY_URL_HINTS = ["/kategori", "/category", "/categories", "/collections",
                       "/urunler", "/urun-listesi", "/c/", "/cat/", "/shop/", "/koleksiyon"]
PRODUCT_URL_HINTS  = ["/urun/", "/product/", "/p/", "/products/", "/detay/", "/dp/"]

# Ürün URL'lerinin sonunda genelde bir sayısal/alfanumerik ID bulunur — örn.
# "...-p-123456", "...-12345.html", "...?pid=123". Bu, platform bağımsız çok
# güçlü bir "bu bir ürün sayfası" sinyalidir.
PRODUCT_ID_SUFFIX_RE = re.compile(r'[-_](p-?)?\d{4,}(\.html?)?/?$', re.IGNORECASE)
# Kategori sayfaları genelde URL'in son segmentinde sayı taşımaz, kelime taşır.
CATEGORY_WORDY_END_RE = re.compile(r'/[a-zçğıöşü\-]+/?$', re.IGNORECASE)


def classify_sitemap_urls(urls):
    """
    URL'leri kategori / ürün / diğer olarak ayırır. Öncelik sırası:
    1. URL sonunda sayısal ID var mı? (en güçlü ürün sinyali, platform bağımsız)
    2. Path'te bilinen ürün hint'i var mı?
    3. Path'te bilinen kategori hint'i var mı?
    4. Hiçbiri yoksa: path derinliğine bak (genelde ürünler daha derin nestlenir)
    """
    category_urls, product_urls, other_urls = [], [], []
    for u in urls:
        path = urlparse(u).path.lower()
        segments = [s for s in path.split("/") if s]

        has_product_hint  = any(h in path for h in PRODUCT_URL_HINTS)
        has_category_hint = any(h in path for h in CATEGORY_URL_HINTS)
        has_numeric_id     = bool(PRODUCT_ID_SUFFIX_RE.search(path))

        if has_numeric_id:
            product_urls.append(u)
        elif has_product_hint and not has_category_hint:
            product_urls.append(u)
        elif has_category_hint and not has_numeric_id:
            category_urls.append(u)
        elif len(segments) >= 3:
            # Derin path (3+ segment) ve hiçbir hint yok — büyük olasılıkla ürün detay sayfası
            product_urls.append(u)
        elif len(segments) in (1, 2):
            # Sığ path (1-2 segment), hint yok — büyük olasılıkla kategori/koleksiyon sayfası
            category_urls.append(u)
        else:
            other_urls.append(u)
    return category_urls, product_urls, other_urls


def discover_sitemaps(domain, progress_cb=None):
    candidates = [f"https://{domain}/sitemap.xml", f"https://{domain}/sitemap_index.xml"]
    try:
        r = requests.get(f"https://{domain}/robots.txt", headers=HEADERS, timeout=10)
        for line in r.text.splitlines():
            if line.lower().startswith("sitemap:"):
                candidates.append(line.split(":", 1)[1].strip())
    except Exception:
        pass
    found = []
    for url in dict.fromkeys(candidates):
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            if r.status_code == 200 and ("<url>" in r.text or "<sitemap>" in r.text):
                found.append(url)
        except Exception:
            pass
    if progress_cb:
        progress_cb(f"{len(found)} sitemap dosyası bulundu.")
    return found


def parse_sitemap_urls(sitemap_url, progress_cb=None, _depth=0):
    """Sitemap XML'ini (ve iç içe sitemap index'lerini) açıp tüm <url><loc> değerlerini döndürür."""
    urls = []
    if _depth > 3:
        return urls
    try:
        r = requests.get(sitemap_url, headers=HEADERS, timeout=20)
        root = ET.fromstring(r.content)
        ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
        child_sitemaps = [loc.text.strip() for loc in root.findall("sm:sitemap/sm:loc", ns)]
        for child in child_sitemaps:
            urls.extend(parse_sitemap_urls(child, progress_cb, _depth + 1))
        urls.extend(loc.text.strip() for loc in root.findall("sm:url/sm:loc", ns))
    except Exception as e:
        if progress_cb:
            progress_cb(f"⚠️ Sitemap parse hatası ({sitemap_url}): {e}")
    return urls


def estimate_site_size(domain, progress_cb=None):
    """
    Hızlı keşif: sitemap'i tarar, kategori/ürün URL sayısını TAHMİN eder.
    Hiçbir sayfa içeriği çekmez — sadece sitemap XML'inden URL sayar.
    Kullanıcıya 'tüm başlıkları çek mi / hızlı tahmin mi' seçimi sunmak için kullanılır.
    """
    sitemaps = discover_sitemaps(domain, progress_cb)
    if not sitemaps:
        return {"category_count": 0, "product_count": 0, "sitemaps_found": 0,
                 "category_urls": [], "product_urls": []}
    all_urls = []
    for sm in sitemaps:
        all_urls.extend(parse_sitemap_urls(sm, progress_cb))
    cat_urls, prod_urls, _ = classify_sitemap_urls(all_urls)
    if progress_cb:
        progress_cb(f"Sitemap taraması: ~{len(cat_urls)} kategori, ~{len(prod_urls)} ürün URL'si bulundu.")
    return {
        "category_count": len(cat_urls),
        "product_count": len(prod_urls),
        "sitemaps_found": len(sitemaps),
        "category_urls": cat_urls,
        "product_urls": prod_urls,
    }


def scrape_page_title_and_count(url, progress_cb=None):
    """Bir kategori sayfasını ziyaret edip başlık + ürün sayısını (sayfadaki metinden) çıkarır."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=8)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        h1 = soup.find("h1")
        title = h1.get_text(strip=True) if h1 else urlparse(url).path.strip("/").split("/")[-1].replace("-", " ").title()
        count = 0
        for pat in [r'(\d+)\s*(ürün|products?|sonuç|results?|kayıt)']:
            m = re.search(pat, r.text, re.IGNORECASE)
            if m:
                count = int(m.group(1)); break
        if not count:
            cards = soup.select(".product-item,.product-card,[data-product-id],article.product,.product")
            count = len(cards)
        handle = urlparse(url).path.strip("/").split("/")[-1]
        return {"title": title, "handle": handle, "url": url, "product_count": count, "source": "sitemap_scrape"}
    except Exception:
        return None


def scrape_product_title(url):
    """Bir ürün sayfasını ziyaret edip sadece başlığını çeker (hızlı, minimal parse)."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=8)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        h1 = soup.find("h1")
        title = h1.get_text(strip=True) if h1 else None
        if not title:
            t = soup.find("title")
            title = t.get_text(strip=True) if t else ""
        return {"title": title, "handle": urlparse(url).path.strip("/").split("/")[-1]}
    except Exception:
        return None


def fetch_generic_categories(domain, category_urls, progress_cb=None, max_categories=150):
    categories = []
    for i, url in enumerate(category_urls[:max_categories]):
        info = scrape_page_title_and_count(url)
        if info:
            categories.append(info)
        if progress_cb and i % 10 == 0:
            progress_cb(f"Kategori sayfaları taranıyor... ({i+1}/{min(len(category_urls), max_categories)})")
        time.sleep(0.15)
    return categories


def fetch_generic_products_full(domain, product_urls, progress_cb=None, max_products=15000):
    """TÜM ürün sayfalarını tek tek ziyaret edip başlık çeker. Yavaş ama tam doğru."""
    products = []
    urls = product_urls[:max_products]
    for i, url in enumerate(urls):
        info = scrape_product_title(url)
        if info and info["title"]:
            products.append(info)
        if progress_cb and i % 25 == 0:
            progress_cb(f"Ürün başlıkları çekiliyor... ({i+1}/{len(urls)})")
        time.sleep(0.2)
    if progress_cb:
        progress_cb(f"Toplam {len(products)} ürün başlığı çekildi.")
    return products


SEARCH_PATH_CANDIDATES = ["/search?q=", "/arama?q=", "/search?query=", "/?s=", "/search/?q="]


def find_working_search_path(domain, progress_cb=None):
    """
    Sitenin hangi arama URL kalıbını kullandığını BİR KERE tespit eder.
    Bilinen bir test kelimesiyle (örn. 'ayakkabı') her kalıbı dener, ilk
    çalışanı döndürür. Sonradan her gap için bu sabit kalıp kullanılır —
    böylece her sorguda 5 farklı path denenmek zorunda kalınmaz.
    """
    test_query = "ayakkabi"
    for path in SEARCH_PATH_CANDIDATES:
        url = f"https://{domain}{path}{test_query}"
        try:
            r = requests.get(url, headers=HEADERS, timeout=8)
            if r.status_code == 200 and len(r.text) > 500:
                if progress_cb:
                    progress_cb(f"Arama endpoint'i bulundu: {path}")
                return path
        except Exception:
            continue
    if progress_cb:
        progress_cb("⚠️ Çalışan bir arama endpoint'i bulunamadı, ürün eşleştirmesi yapılamayacak.")
    return None


def try_search_count(domain, query, search_path=None, progress_cb=None):
    """Sonuç sayısını (yaklaşık) okur. search_path verilmişse SADECE onu dener (hızlı)."""
    paths = [search_path] if search_path else SEARCH_PATH_CANDIDATES
    for path in paths:
        url = f"https://{domain}{path}{query.replace(' ', '+')}"
        try:
            r = requests.get(url, headers=HEADERS, timeout=8)
            if r.status_code != 200:
                continue
            for pat in [r'(\d+)\s*(ürün|sonuç|results?|products?)\s*bulundu',
                        r'(\d+)\s*(ürün|sonuç|results?|products?)']:
                m = re.search(pat, r.text, re.IGNORECASE)
                if m:
                    return int(m.group(1))
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.select(".product-item,.product-card,[data-product-id],article.product,.product")
            if cards:
                return len(cards)
            return 0
        except Exception:
            continue
    return None


def fetch_generic_products_via_search(domain, queries, progress_cb=None):
    """
    Tüm ürünleri çekmek yerine, gap adaylarının kelimelerini arama endpoint'inden
    sorgulayıp sonuç sayısını okur. queries: aranacak kelime öbeklerinin listesi.
    Döner: {query: result_count}
    """
    results = {}
    for i, q in enumerate(queries):
        count = try_search_count(domain, q, progress_cb)
        results[q] = count or 0
        if progress_cb and i % 10 == 0:
            progress_cb(f"Arama ile ürün sayısı kontrol ediliyor... ({i+1}/{len(queries)})")
        time.sleep(0.3)
    return results


def fetch_nav_tree(domain, progress_cb=None):
    """
    Ana sayfadaki navigasyon menüsünden Kadın > Çanta > Deri Çanta gibi
    hiyerarşiyi çıkarır. Shopify temaları çok farklı HTML yapıları kullanabilir
    (mega-menu'lerde <li> ile alt <ul> arasına <div> girebilir), bu yüzden
    esnek bir arama stratejisi kullanılır. Bulamazsak boş döner (flat liste
    fallback'i devam eder).

    Dönen yapı:
    {
        "Kadın": {"handle": "kadin", "children": {
            "Çanta": {"handle": "kadin-canta", "children": {
                "Deri Çanta": {"handle": "deri-canta", "children": {}}
            }}
        }}
    }
    """
    try:
        r = requests.get(f"https://{domain}/", headers=HEADERS, timeout=20)
        if r.status_code != 200:
            if progress_cb:
                progress_cb(f"⚠️ Ana sayfa HTTP {r.status_code} — navigasyon menüsü atlanıyor.")
            return {}
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        if progress_cb:
            progress_cb(f"⚠️ Ana sayfa çekme hatası: {e}")
        return {}

    def handle_from_href(href):
        m = re.search(r"/collections/([a-z0-9\-]+)", href or "", re.IGNORECASE)
        return m.group(1) if m else None

    def nearest_child_ul(li_tag):
        """<li> içinde, ara div/span'lara bakmadan en yakın <ul>'u bulur (li'nin kendi <ul>'u değilse)."""
        for child in li_tag.find_all("ul"):
            return child   # ilk bulunan iç <ul>, derinlik önemli değil — mega menu'lerde tek seviye olur
        return None

    def parse_list(ul_tag, depth=0):
        """Bir <ul> içindeki <li>'leri (recursive=False ile direkt çocuklar) gezer."""
        tree = {}
        if not ul_tag or depth > 3:
            return tree
        direct_lis = ul_tag.find_all("li", recursive=False)
        if not direct_lis:
            # Bazı temalarda <ul> doğrudan <li> içermez, class bazlı menu item div'leri olur
            direct_lis = ul_tag.find_all(["li"], recursive=True)
        for li in direct_lis:
            a = li.find("a")
            if not a:
                continue
            name = a.get_text(strip=True)
            href = a.get("href", "")
            handle = handle_from_href(href)
            if not name or len(name) > 40 or len(name) < 2:
                continue
            sub_ul = nearest_child_ul(li)
            children = parse_list(sub_ul, depth + 1) if sub_ul else {}
            if handle or children:
                tree.setdefault(name, {"handle": handle, "children": children})
        return tree

    def find_best_nav_candidate():
        """Birden fazla aday dener, en çok /collections/ linki içereni seçer."""
        candidates = []
        for tag in soup.find_all(["nav", "header"]):
            candidates.append(tag)
        for cls_pattern in ["main-nav", "navigation", "menu", "site-nav", "header-nav", "mega-menu"]:
            found = soup.find_all(class_=re.compile(cls_pattern, re.I))
            candidates.extend(found)
        if not candidates:
            candidates = [soup]   # son çare: tüm body'de ara

        best, best_count = None, 0
        for cand in candidates:
            count = len(cand.find_all("a", href=re.compile(r"/collections/")))
            if count > best_count:
                best, best_count = cand, count
        return best, best_count

    nav, link_count = find_best_nav_candidate()
    if not nav or link_count < 2:
        if progress_cb:
            progress_cb("⚠️ Navigasyon menüsü bulunamadı, düz kategori listesi kullanılacak.")
        return {}

    # Nav içindeki en üst seviye <ul>'ları dene — birden fazla olabilir (örn. ayrı sol/sağ menu)
    top_uls = nav.find_all("ul", recursive=False) or nav.find_all("ul")
    tree = {}
    for ul in top_uls[:3]:
        partial = parse_list(ul, 0)
        tree.update(partial)

    if not tree:
        tree = parse_list(nav, 0)

    if progress_cb:
        total_nodes = _count_tree_nodes(tree)
        progress_cb(f"Navigasyon menüsünden {total_nodes} kategori düğümü çıkarıldı ({link_count} link tespit edildi).")
    return tree


def _count_tree_nodes(tree):
    count = len(tree)
    for node in tree.values():
        count += _count_tree_nodes(node.get("children", {}))
    return count


def flatten_nav_tree(tree, parent_path=None):
    """
    Nav ağacını düz bir handle->path eşlemesine çevirir.
    Örn: {"kadin-canta": ["Kadın", "Çanta"], "deri-canta": ["Kadın", "Çanta", "Deri Çanta"]}
    """
    parent_path = parent_path or []
    mapping = {}
    for name, node in tree.items():
        path = parent_path + [name]
        if node.get("handle"):
            mapping[node["handle"]] = path
        mapping.update(flatten_nav_tree(node.get("children", {}), path))
    return mapping


def top_level_groups(tree):
    """Nav ağacının en üst seviyesini döndürür — örn. ['Kadın', 'Erkek', 'Ayakkabı', ...]."""
    return list(tree.keys())


# Nav menüsü bulunamadığında, kategori isimlerinin İLK kelimesinden grup çıkarmak için
# kullanılan bilinen önekler (marka bağımsız, herhangi bir Türkçe e-ticaret sitesinde geçerli).
GENDER_PREFIX_HINTS = {"kadin": "Kadın", "erkek": "Erkek", "cocuk": "Çocuk", "bebek": "Bebek", "unisex": "Unisex"}


def build_fallback_groups(categories):
    """
    Nav menüsü bulunamadığında kategori isimlerinden basit bir gruplama çıkarır.
    'Kadın Bot', 'Kadın Ayakkabı' -> grup 'Kadın'; eşleşmeyenler kendi başlığını grup sayar.
    Her kategoriye nav_path atar (zaten yoksa).
    """
    for cat in categories:
        if cat.get("nav_path") and cat["nav_path"] != [cat["title"]]:
            continue   # zaten gerçek nav_path var, dokunma
        first_word = _norm(cat["title"]).split()[0] if cat["title"] else ""
        group = GENDER_PREFIX_HINTS.get(first_word)
        cat["nav_path"] = [group, cat["title"]] if group else [cat["title"]]
    return categories


def top_level_groups_from_categories(categories):
    """Kategorilerin nav_path'lerinden (gerçek nav ya da fallback) benzersiz üst grupları çıkarır."""
    groups, seen = [], set()
    for cat in categories:
        g = cat.get("nav_path", [cat["title"]])[0]
        if g and g not in seen and g != cat["title"]:
            seen.add(g); groups.append(g)
    return groups


def fetch_site_data(domain, progress_cb=None, force_shopify=False, crawl_mode="full_crawl",
                     category_urls=None, product_urls=None):
    """
    crawl_mode:
      - "full_crawl": (sadece generic platform) tüm ürün sayfalarını tek tek ziyaret eder — yavaş, en doğru
      - "search_estimate": (sadece generic platform) ürün başlığı çekmez, gap analizinde arama
        endpoint'i üzerinden sayım yapılacak şekilde boş ürün listesiyle devam eder
    category_urls / product_urls: estimate_site_size'dan önceden bulunmuşsa tekrar taramamak için verilir.
    """
    platform = "shopify" if force_shopify else detect_platform(domain, progress_cb)

    if platform == "shopify":
        if progress_cb:
            progress_cb("Platform: SHOPIFY" + (" (manuel seçildi)" if force_shopify else ""))
        nav_tree = fetch_nav_tree(domain, progress_cb)
        handle_to_path = flatten_nav_tree(nav_tree) if nav_tree else {}

        collections = fetch_all_collections(domain, progress_cb)
        real_cats, noise_cats = build_categories(collections, progress_cb)

        for cat in real_cats:
            path = handle_to_path.get(cat["handle"])
            cat["nav_path"] = path or [cat["title"]]
        if not nav_tree:
            real_cats = build_fallback_groups(real_cats)
            if progress_cb:
                fb_groups = top_level_groups_from_categories(real_cats)
                progress_cb(f"Nav menü bulunamadı, kategori isimlerinden {len(fb_groups)} grup tahmin edildi.")

        real_cats = fetch_category_product_counts(domain, real_cats, progress_cb)
        products = fetch_all_products(domain, progress_cb)
        return real_cats, noise_cats, products, nav_tree

    # ── Genel platform (Shopify değil) ──
    if progress_cb:
        progress_cb("Platform: GENEL (sitemap tabanlı tarama)")

    nav_tree = fetch_nav_tree(domain, progress_cb)
    handle_to_path = flatten_nav_tree(nav_tree) if nav_tree else {}

    if category_urls is None or product_urls is None:
        size_info = estimate_site_size(domain, progress_cb)
        category_urls = size_info["category_urls"]
        product_urls = size_info["product_urls"]

    raw_cats = fetch_generic_categories(domain, category_urls, progress_cb)
    real_cats, noise_cats = build_categories(raw_cats, progress_cb)
    for cat in real_cats:
        path = handle_to_path.get(cat["handle"])
        cat["nav_path"] = path or [cat["title"]]
    if not nav_tree:
        real_cats = build_fallback_groups(real_cats)
        if progress_cb:
            fb_groups = top_level_groups_from_categories(real_cats)
            progress_cb(f"Nav menü bulunamadı, kategori isimlerinden {len(fb_groups)} grup tahmin edildi.")

    if crawl_mode == "full_crawl":
        products = fetch_generic_products_full(domain, product_urls, progress_cb)
    else:
        # search_estimate modunda ürün başlıkları çekilmez; gap analizi sırasında
        # arama endpoint'i üzerinden canlı sayım yapılır (run_gap_analysis -> search_fallback)
        products = []

    return real_cats, noise_cats, products, nav_tree


# ══════════════════════════════════════════════════════════════
# 2. DATAFORSEO
# ══════════════════════════════════════════════════════════════

AUTH = (DFS_LOGIN, DFS_PASSWORD)
BASE = "https://api.dataforseo.com/v3"

def _dfs_post(endpoint, payload, progress_cb=None):
    try:
        r    = requests.post(f"{BASE}/{endpoint}", auth=AUTH, json=payload, timeout=60)
        data = r.json()
        if data.get("status_code") != 20000:
            if progress_cb:
                progress_cb(f"⚠️ DataForSEO hatası: {data.get('status_code')} {data.get('status_message')}")
            return None
        task = data.get("tasks", [{}])[0]
        if task.get("status_code") != 20000:
            if progress_cb:
                progress_cb(f"⚠️ DataForSEO task hatası: {task.get('status_code')} {task.get('status_message')}")
            return None
        return data
    except Exception as e:
        if progress_cb:
            progress_cb(f"⚠️ İstek hatası: {e}")
        return None

def fetch_domain_keywords(domain, progress_cb=None):
    if progress_cb:
        progress_cb("DataForSEO'dan keyword'ler çekiliyor...")
    payload = [{"target": domain, "language_code": "tr", "location_code": 2792,
                "limit": 1000, "filters": ["keyword_data.keyword_info.search_volume", ">", 0],
                "order_by": ["keyword_data.keyword_info.search_volume,desc"]}]
    data = _dfs_post("dataforseo_labs/google/ranked_keywords/live", payload, progress_cb)
    if not data:
        return []
    keywords = []
    try:
        result = data["tasks"][0].get("result")
        if not result or not result[0].get("items"):
            if progress_cb:
                progress_cb("⚠️ DataForSEO sonuç döndü ama keyword listesi boş.")
            return []
        for item in result[0]["items"]:
            kw   = item.get("keyword_data", {})
            info = kw.get("keyword_info", {})
            serp = item.get("ranked_serp_element", {}).get("serp_element", {})
            keywords.append({"keyword": kw.get("keyword",""), "search_volume": info.get("search_volume",0),
                             "cpc": info.get("cpc",0), "competition": info.get("competition",0),
                             "position": serp.get("rank_group",0), "url": serp.get("url","")})
        if progress_cb:
            progress_cb(f"{len(keywords)} keyword çekildi.")
    except Exception as e:
        if progress_cb:
            progress_cb(f"⚠️ Parse hatası: {e}")
    return keywords


# ══════════════════════════════════════════════════════════════
# 3. ANALYZER — keyword × kategori × ürün eşleştirmesi
# ══════════════════════════════════════════════════════════════

STOP = {"ve","ile","icin","bir","bu","da","de","mi","mu","the","and","for","in","of","a","an","on","at","to"}
TR_NORM = str.maketrans("ığüşöçİĞÜŞÖÇ","igusocIGUSOC")
MODIFIERS = set([
    "siyah","beyaz","kirmizi","mavi","yesil","sari","mor","pembe","turuncu","gri",
    "kahverengi","bej","lacivert","bordo","krem","ekru","haki","black","white","red",
    "blue","green","yellow","grey","brown","beige","navy","gumus","altin",
    "deri","suet","tokali","bagcikli","fermuarli","platform","topuklu","duz","kalin",
    "ince","uzun","kisa","maxi","mini","midi","spor","klasik","casual","formal",
    "gunluk","yazlik","kislik","sik","rahat",
])

def _norm(text):
    return re.sub(r"[^a-z0-9\s]"," ", text.lower().translate(TR_NORM)).strip()

def _tokens(text):
    return set(_norm(text).split()) - STOP

def _matches_category(keyword, cat):
    return bool(_tokens(keyword) & (_tokens(cat["title"]) | _tokens(cat["handle"])))

def count_matching_products(query_tokens, products):
    if not query_tokens:
        return 0
    count = 0
    for p in products:
        p_tokens = _tokens(p["title"])
        if query_tokens & p_tokens and len(query_tokens & p_tokens) >= len(query_tokens):
            count += 1
    return count

NON_CATEGORY_PATTERNS = [
    # Mağaza / lokasyon / AVM isimleri — marka bağımsız
    r'\bavm\b', r'\boutlet\b', r'\bmagaza\w*\b', r'\bmağaza\w*\b',
    r'\bşube\w*\b', r'\bsube\w*\b', r'\bmarina\b', r'\bpower outlet\b', r'\bplaza\b',
    # "X ne demek / nedir" gibi tanım soruları — ürün kategorisi değil, içerik/blog sorusu
    r'\bne demek\b', r'\bnedir\b', r'\bne anlama gelir\b',
    # "boykot", "şikayet", haber/yorum tarzı sorgular — e-ticaret kategorisi değil
    r'\bboykot\b', r'\bşikayet\w*\b', r'\bsikayet\w*\b', r'\biade\b', r'\biflas\b',
    # Soyut/şirket-bilgisi sorguları
    r'\bcompany\b', r'\bfor sale\b', r'\bnasıl\b', r'\bne kadar\b', r'\bfiyatı ne\b',
]
NON_CATEGORY_RE = re.compile("|".join(NON_CATEGORY_PATTERNS), re.IGNORECASE)

# Bilinen 3. parti / rakip ayakkabı-giyim markaları — herhangi bir e-ticaret sitesi için
# bu markaların adını arayan keyword'ler o sitenin kendi kategori önerisi OLAMAZ
# (rakip marka taraması ya da zaten var olan "marka filtresi" ihtiyacıdır, otomatik
# kategori önerisi olarak yanıltıcıdır). Marka-bağımsız genel bir kural.
THIRD_PARTY_BRANDS = {
    "nike","adidas","puma","reebok","skechers","caterpillar","cat","new balance",
    "converse","vans","crocs","clarks","timberland","ecco","geox","lacoste",
    "tommy","calvin klein","levis","mango","zara","h&m","pull bear","bershka",
    "north face","columbia","under armour","asics","fila","kappa","lotto",
}

def _contains_third_party_brand(tokens):
    joined = " ".join(tokens)
    return any(b in joined for b in THIRD_PARTY_BRANDS)

# Türkiye'deki büyük şehir/ilçe isimleri — "X Fabrika", "X Sanayi", "X Fotoğrafları" gibi
# yerel/coğrafi sorgular ürün kategorisi olamaz (herhangi bir marka için).
# NOT: _norm() Türkçe karakterleri ASCII'ye çevirir (ş→s, ı→i, ö→o, ü→u, ç→c, ğ→g),
# bu set de aynı normalize edilmiş haliyle tutulur.
TR_PLACE_NAMES = {
    "istanbul","ankara","izmir","bursa","antalya","adana","konya","mersin",
    "kocaeli","izmit","eskisehir","gaziantep","kayseri","samsun",
    "denizli","sanliurfa","adapazari","malatya",
    "kahramanmaras","erzurum","diyarbakir",
    "sivas","balikesir","manisa","tarsus","trabzon","alanya",
    "edremit","aydin","mugla","van","batman","elazig",
}
ABSTRACT_PLACE_SUFFIXES = {"fabrika","sanayi","fotograflari","sitesi","depo","yapi"}

# Bu kelimeler tek başına da (şehir adı olmadan) çok büyük olasılıkla bir ürün
# kategorisi değil, bir kurum/site/depo/firma adını işaret eder — marka bağımsız.
STANDALONE_NON_CATEGORY_SUFFIXES = {"sitesi","depo","fabrika","plaza","center","sirketi","şirketi"}

def _looks_like_place_query(tokens):
    """'<şehir> <soyut kelime>' kalıbı — örn. 'eskişehir fabrika', 'adana fotoğrafları'. tokens zaten _norm() ile ASCII'ye çevrilmiş olmalı."""
    if any(t in TR_PLACE_NAMES for t in tokens) and any(t in ABSTRACT_PLACE_SUFFIXES for t in tokens):
        return True
    # Şehir adı olmasa da "X Sitesi", "X Depo", "X Plaza" gibi kurumsal/lokasyon kalıpları
    if any(t in STANDALONE_NON_CATEGORY_SUFFIXES for t in tokens):
        return True
    # "<özel isim> <şehir adı>" kalıbı — örn. "espark eskişehir", "deepo antalya".
    # Ürün kategorisi isimleri neredeyse hiçbir zaman bir şehir adıyla bitmez.
    if len(tokens) == 2 and tokens[1] in TR_PLACE_NAMES:
        return True
    return False


def run_gap_analysis(keywords, categories, products, progress_cb=None, domain=None, use_search_fallback=False):
    if progress_cb:
        progress_cb("Gap analizi yapılıyor...")

    matched, orphans = [], []
    groups = defaultdict(lambda: {"keywords": [], "total_volume": 0, "parent_cat": None})

    for kw in keywords:
        word, vol = kw.get("keyword", ""), kw.get("search_volume", 0)
        if vol < 100:
            continue
        if NON_CATEGORY_RE.search(word):
            orphans.append(kw); continue

        norm_word = _norm(word)
        tokens = norm_word.split()
        if _looks_like_place_query(tokens):
            orphans.append(kw); continue
        if _contains_third_party_brand(tokens):
            orphans.append(kw); continue

        # En iyi eşleşen mevcut kategoriyi bul (en çok ortak token'a sahip olan)
        best_cat, best_overlap = None, 0
        for c in categories:
            cat_tokens = _tokens(c["title"]) | _tokens(c["handle"])
            overlap = len(_tokens(word) & cat_tokens)
            if overlap > best_overlap:
                best_overlap, best_cat = overlap, c

        kw_tokens = _tokens(word)
        cat_tokens = (_tokens(best_cat["title"]) | _tokens(best_cat["handle"])) if best_cat else set()
        extra_tokens = kw_tokens - cat_tokens   # keyword'de olup kategori isminde olmayan kelimeler

        if best_cat and not extra_tokens:
            # Keyword zaten kategori ismiyle tam örtüşüyor → gerçekten kategoriye ait, gap değil
            matched.append(kw); continue

        if best_cat and extra_tokens and best_overlap > 0:
            # Kısmi eşleşme: "hasır çanta" ~ "Çanta" kategorisi + "hasır" ek kelime → alt kategori adayı
            if len(tokens) < 2 and norm_word not in MODIFIERS and not extra_tokens & MODIFIERS:
                orphans.append(kw); continue
            key = norm_word
            groups[key]["keywords"].append(kw)
            groups[key]["total_volume"] += vol
            groups[key]["parent_cat"] = best_cat
            continue

        # Hiçbir kategoriyle örtüşmedi — tamamen yeni bir kategori adayı
        if len(tokens) < 2 and norm_word not in MODIFIERS:
            orphans.append(kw); continue
        key = norm_word
        groups[key]["keywords"].append(kw)
        groups[key]["total_volume"] += vol

    # ── 1. geçiş: tüm gap'leri oluştur (arama yapmadan, products varsa onunla say) ──
    raw_gaps = []
    for key, group in groups.items():
        if not group["keywords"]:
            continue
        top = max(group["keywords"], key=lambda x: x["search_volume"])
        vol = group["total_volume"]
        parent_cat = group.get("parent_cat")
        query_tokens = _tokens(top["keyword"])
        raw_gaps.append({
            "key": key, "group": group, "top": top, "vol": vol,
            "parent_cat": parent_cat, "query_tokens": query_tokens,
        })
    raw_gaps.sort(key=lambda x: x["vol"], reverse=True)

    # ── Arama-fallback modunda: SADECE en yüksek hacimli N gap için canlı sorgu yap ──
    # (yüzlerce gap için tek tek arama yapmak saatler sürebilir; en değerli adaylara odaklanılır)
    MAX_SEARCH_QUERIES = 80
    search_path = None
    if use_search_fallback and domain and raw_gaps:
        search_path = find_working_search_path(domain, progress_cb)

    gaps = []
    for i, rg in enumerate(raw_gaps):
        top, vol, parent_cat, query_tokens = rg["top"], rg["vol"], rg["parent_cat"], rg["query_tokens"]
        label = top["keyword"].title()

        if use_search_fallback and domain and search_path and i < MAX_SEARCH_QUERIES:
            real_product_count = try_search_count(domain, top["keyword"], search_path, progress_cb) or 0
            if progress_cb and i % 10 == 0:
                progress_cb(f"Ürün eşleştirmesi kontrol ediliyor... ({i+1}/{min(len(raw_gaps), MAX_SEARCH_QUERIES)})")
        elif use_search_fallback and (not search_path or i >= MAX_SEARCH_QUERIES):
            real_product_count = 0   # arama yapılamadı / limit aşıldı — bilinmiyor sayılır
        else:
            real_product_count = count_matching_products(query_tokens, products)

        if real_product_count > 0:
            status = "opportunity"
        elif vol >= 500:
            status = "high"
        else:
            status = "none"

        nav_group = (parent_cat.get("nav_path", [parent_cat["title"]])[0] if parent_cat
                     else (top.get("nav_group") or "Diğer"))

        gaps.append({
            "suggested_category":     label,
            "modifier":                next((t for t in query_tokens if t in MODIFIERS), ""),
            "base_category":           parent_cat["title"] if parent_cat else "Belirsiz",
            "nav_group":               nav_group,
            "keyword_count":           len(rg["group"]["keywords"]),
            "total_search_volume":     vol,
            "top_keyword":             top["keyword"],
            "top_kw_volume":           top["search_volume"],
            "matching_product_count":  real_product_count,
            "status":                  status,
            "sample_keywords":         ", ".join(k["keyword"] for k in sorted(
                rg["group"]["keywords"], key=lambda x: x["search_volume"], reverse=True)[:5]),
        })
    gaps.sort(key=lambda x: x["total_search_volume"], reverse=True)
    return {"matched": matched, "gaps": gaps, "orphans": orphans}


# ══════════════════════════════════════════════════════════════
# 4. CHARTS
# ══════════════════════════════════════════════════════════════

def build_top_categories_table(categories, matched_keywords, limit=10):
    """Mevcut kategorilere eşleşen keyword hacimlerini toplayıp en yüksek performanslı kategorileri sıralar."""
    vol_by_cat = defaultdict(int)
    kw_count_by_cat = defaultdict(int)
    for kw in matched_keywords:
        word = kw.get("keyword", "")
        vol = kw.get("search_volume", 0)
        for c in categories:
            if _matches_category(word, c):
                vol_by_cat[c["title"]] += vol
                kw_count_by_cat[c["title"]] += 1
                break
    rows = []
    for cat in categories:
        title = cat["title"]
        if vol_by_cat.get(title, 0) > 0:
            rows.append({
                "title": title,
                "product_count": cat["product_count"] or 0,
                "total_volume": vol_by_cat[title],
                "keyword_count": kw_count_by_cat[title],
            })
    rows.sort(key=lambda r: r["total_volume"], reverse=True)
    return rows[:limit]


STATUS_LABEL = {"opportunity": "ÜRÜN VAR · SAYFA YOK", "high": "HACİM VAR · ÜRÜN YOK", "none": "DÜŞÜK ÖNCELİK"}
STATUS_BADGE_CLASS = {"opportunity": "badge-opportunity", "high": "badge-high", "none": "badge-none"}
STATUS_COLOR = {"opportunity": "#1a9b5c", "high": "#d97a1f", "none": "#c23d3d"}


def build_treemap(categories, gaps, max_categories=20, max_gaps=10):
    top_cats = sorted(categories, key=lambda c: c["product_count"] or 0, reverse=True)[:max_categories]
    top_gaps = gaps[:max_gaps]
    labels, parents, values, colors, texts = ["Site"], [""], [0], ["#eceef3"], [""]

    for cat in top_cats:
        labels.append(cat["title"]); parents.append("Site")
        values.append(max(cat["product_count"] or 0, 5))
        colors.append("#cdeedb")
        texts.append(f"Ürün: {cat['product_count']}")

    for gap in top_gaps:
        labels.append(f"+ {gap['suggested_category']}"); parents.append("Site")
        values.append(max(gap["total_search_volume"] // 5, 30))
        colors.append({"opportunity":"#bfe8d2","high":"#fad9b3","none":"#f7c4c4"}[gap["status"]])
        texts.append(f"Hacim: {gap['total_search_volume']:,} · Eşleşen ürün: {gap['matching_product_count']}")

    fig = go.Figure(go.Treemap(
        labels=labels, parents=parents, values=values, customdata=texts,
        hovertemplate="<b>%{label}</b><br>%{customdata}<extra></extra>",
        marker=dict(colors=colors, line=dict(width=2, color="#fff")),
        textfont=dict(family="Inter", size=13, color="#14161f"),
        pathbar=dict(visible=False),
    ))
    fig.update_layout(margin=dict(l=0,r=0,t=0,b=0), paper_bgcolor="#fff", plot_bgcolor="#fff", height=420)
    return fig


def build_volume_bar(gaps):
    if not gaps:
        return None
    top = gaps[:15]
    fig = go.Figure(go.Bar(
        x=[g["total_search_volume"] for g in top],
        y=[g["suggested_category"] for g in top],
        orientation="h",
        marker_color=[STATUS_COLOR[g["status"]] for g in top],
        text=[f"{g['total_search_volume']:,}" for g in top],
        textposition="outside",
        textfont=dict(family="JetBrains Mono", size=11, color="#4a4e5c"),
        hovertemplate="<b>%{y}</b><br>Hacim: %{x:,}<extra></extra>",
    ))
    fig.update_layout(
        margin=dict(l=0,r=50,t=0,b=0), paper_bgcolor="#fff", plot_bgcolor="#fff",
        height=max(280, len(top)*30),
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(showgrid=False, color="#4a4e5c", tickfont=dict(family="Inter", size=12)),
    )
    return fig


# ══════════════════════════════════════════════════════════════
# 5. EXCEL EXPORT
# ══════════════════════════════════════════════════════════════

def _hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, r, c, v, bg=None, bold=False, align="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    s = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=s, right=s, top=s, bottom=s)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def generate_excel(gaps, categories, noise_categories, keywords, domain):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    status_label = {"opportunity": "URUN VAR - SAYFA YOK", "high": "HACIM VAR - URUN YOK", "none": "DUSUK ONCELIK"}
    status_color = {"opportunity": "C6EFCE", "high": "FFE2B3", "none": "FFC7CE"}

    ws = wb.create_sheet("Kategori Onerileri")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    ws["A1"] = f"SEO Kategori Gap Analizi - {domain}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Olusturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    hdrs = ["Onerilen Kategori","Modifier","Baz Kategori","Keyword Sayisi","Toplam Hacim",
            "En Yuksek KW","En Yuksek KW Hacmi","Eslesen Urun Sayisi","Durum","Ornek Keywordler"]
    widths = [28,15,22,14,14,28,18,16,22,50]
    for c,(h,w) in enumerate(zip(hdrs,widths),1):
        _hdr(ws,3,c,h); ws.column_dimensions[get_column_letter(c)].width = w
    for i,g in enumerate(gaps,4):
        bg = status_color.get(g["status"])
        row = [g["suggested_category"],g["modifier"],g["base_category"],g["keyword_count"],
               g["total_search_volume"],g["top_keyword"],g["top_kw_volume"],g["matching_product_count"],
               status_label.get(g["status"]), g["sample_keywords"]]
        for c,v in enumerate(row,1):
            _cell(ws,i,c,v,bg=bg,bold=(c==9),align="right" if isinstance(v,int) else "left")
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("Mevcut Kategoriler")
    ws2.sheet_view.showGridLines = False
    for c,(h,w) in enumerate(zip(["Kategori","Handle","Urun Sayisi"],[28,25,14]),1):
        _hdr(ws2,1,c,h); ws2.column_dimensions[get_column_letter(c)].width = w
    for i,cat in enumerate(categories,2):
        bg = "F2F2F2" if i%2==0 else None
        for c,v in enumerate([cat["title"],cat["handle"],cat["product_count"]],1):
            _cell(ws2,i,c,v,bg=bg,align="right" if isinstance(v,int) else "left")
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Filtrelenenler")
    ws3.sheet_view.showGridLines = False
    for c,(h,w) in enumerate(zip(["Koleksiyon","Handle"],[28,25]),1):
        _hdr(ws3,1,c,h); ws3.column_dimensions[get_column_letter(c)].width = w
    for i,cat in enumerate(noise_categories,2):
        bg = "F2F2F2" if i%2==0 else None
        for c,v in enumerate([cat["title"],cat["handle"]],1):
            _cell(ws3,i,c,v,bg=bg)
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("Keyword Listesi")
    ws4.sheet_view.showGridLines = False
    for c,(h,w) in enumerate(zip(["Keyword","Hacim","CPC ($)","Rekabet","Pozisyon","URL"],[35,13,10,12,12,55]),1):
        _hdr(ws4,1,c,h); ws4.column_dimensions[get_column_letter(c)].width = w
    for i,kw in enumerate(keywords,2):
        bg = "F2F2F2" if i%2==0 else None
        for c,v in enumerate([kw["keyword"],kw["search_volume"],kw["cpc"],kw["competition"],kw["position"],kw["url"]],1):
            _cell(ws4,i,c,v,bg=bg,align="right" if isinstance(v,(int,float)) else "left")
    ws4.freeze_panes = "A2"

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════

st.markdown("""
<div class="app-header">
    <h1>SEO Gap Analyzer</h1>
    <p>Kategori ağacını analiz et, hacmi olan ama sayfası olmayan fırsatları bul.</p>
</div>
""", unsafe_allow_html=True)

col_input, col_btn = st.columns([5, 1])
with col_input:
    domain_input = st.text_input("", placeholder="derimod.com.tr", label_visibility="collapsed")
with col_btn:
    discover_btn = st.button("Analiz Et")

force_shopify = st.checkbox(
    "Bu site Shopify (platform tespiti rate-limit'e çarparsa bunu işaretleyip tekrar deneyin)",
    value=False,
)

if "results" not in st.session_state:
    st.session_state.results = None
if "discovery" not in st.session_state:
    st.session_state.discovery = None

def cb_factory(log_lines, status):
    def cb(msg):
        log_lines.append(msg)
        html = "".join(f"<p style='color:{'#d23c3c' if '⚠' in l else '#767a8a'};font-size:0.8rem;margin:2px 0'>{l}</p>" for l in log_lines[-6:])
        status.markdown(html, unsafe_allow_html=True)
    return cb

# ── 1. AŞAMA: Keşif ──
if discover_btn and domain_input:
    domain = domain_input.strip().replace("https://", "").replace("http://", "").strip("/")
    status = st.empty()
    log_lines = []
    cb = cb_factory(log_lines, status)

    platform = "shopify" if force_shopify else detect_platform(domain, cb)

    if platform == "shopify":
        # Shopify hızlı API yolu kullandığı için keşif adımına gerek yok, direkt analiz başlar
        st.session_state.discovery = {"domain": domain, "platform": "shopify", "force_shopify": force_shopify}
    else:
        cb("Site büyüklüğü tahmin ediliyor (sitemap taranıyor)...")
        size_info = estimate_site_size(domain, cb)
        st.session_state.discovery = {
            "domain": domain, "platform": "generic",
            "size_info": size_info,
        }
    status.empty()

# ── Genel platform için: kullanıcıya strateji seçimi sun ──
disc = st.session_state.discovery
if disc and disc.get("platform") == "generic" and "crawl_mode" not in disc:
    size_info = disc["size_info"]
    p_count = size_info["product_count"]
    c_count = size_info["category_count"]

    est_cat_minutes  = round((min(c_count, 150) * 0.6) / 60, 1) or 0.1   # kategori taraması her iki modda da çalışır
    est_full_minutes = round(est_cat_minutes + (p_count * 0.5) / 60)
    est_search_minutes = round(est_cat_minutes + 1)   # gap'ler belirlendikten sonra sadece o kelimeler aranır

    st.markdown(f"""
    <div class="warn-box" style="background:#eaf1fd;border-color:#bcd4f5;color:#2a4d7a">
        Sitemap'te yaklaşık <b>{c_count} kategori</b> ve <b>{p_count:,} ürün</b> URL'si bulundu.
        {"İlk 150 kategori taranacak. " if c_count > 150 else ""}Nasıl ilerleyelim?
    </div>
    """, unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown(f"**Tüm ürün başlıklarını çek**  \nEn doğru sonuç. Tahmini süre: ~{est_full_minutes} dakika.")
        if st.button("Bu yöntemle ilerle", key="full_crawl_btn"):
            disc["crawl_mode"] = "full_crawl"
            st.rerun()
    with col_b:
        st.markdown(f"**Arama endpoint'i ile hızlı tahmin**  \nDaha az kesin ama çok daha hızlı. Tahmini süre: ~{est_search_minutes} dakika.")
        if st.button("Bu yöntemle ilerle", key="search_estimate_btn"):
            disc["crawl_mode"] = "search_estimate"
            st.rerun()

# ── 2. AŞAMA: Asıl analiz ──
ready_shopify = disc and disc.get("platform") == "shopify"
ready_generic = disc and disc.get("platform") == "generic" and "crawl_mode" in disc

if ready_shopify or ready_generic:
    domain = disc["domain"]
    status = st.empty()
    prog   = st.progress(0)
    log_lines = []
    cb = cb_factory(log_lines, status)

    cb("Site verisi çekiliyor (büyük sitelerde birkaç dakika sürebilir)...")
    prog.progress(5)

    if ready_shopify:
        real_cats, noise_cats, products, nav_tree = fetch_site_data(
            domain, cb, force_shopify=disc.get("force_shopify", False)
        )
        use_search_fallback = False
    else:
        size_info = disc["size_info"]
        real_cats, noise_cats, products, nav_tree = fetch_site_data(
            domain, cb, crawl_mode=disc["crawl_mode"],
            category_urls=size_info["category_urls"], product_urls=size_info["product_urls"],
        )
        use_search_fallback = (disc["crawl_mode"] == "search_estimate")

    prog.progress(55)
    keywords = fetch_domain_keywords(domain, cb)

    prog.progress(80)
    results = run_gap_analysis(keywords, real_cats, products, cb, domain=domain, use_search_fallback=use_search_fallback)
    results.update({"categories": real_cats, "noise_categories": noise_cats,
                     "products": products, "keywords": keywords, "domain": domain,
                     "logs": log_lines, "nav_tree": nav_tree})
    prog.progress(100)

    status.empty(); prog.empty()
    st.session_state.results = results
    st.session_state.discovery = None   # bir sonraki analiz için sıfırla
    st.session_state.results = results

if st.session_state.results:
    r = st.session_state.results
    gaps_all, cats, noise_cats = r["gaps"], r["categories"], r.get("noise_categories", [])
    products, kws, dom = r.get("products", []), r["keywords"], r["domain"]
    nav_tree = r.get("nav_tree", {})

    warnings = [l for l in r.get("logs", []) if "⚠" in l]
    if warnings:
        st.markdown("<div class='warn-box'>" + "<br>".join(warnings) + "</div>", unsafe_allow_html=True)

    # ── Üst seviye grup navigasyonu (Kadın / Erkek / Ayakkabı / Çanta ...) ──
    nav_groups = top_level_groups(nav_tree) if nav_tree else top_level_groups_from_categories(cats)
    selected_group = "Tümü"
    if nav_groups:
        st.markdown('<div class="section-title">Kategori Grubu Seç</div>', unsafe_allow_html=True)
        selected_group = st.radio(
            "", ["Tümü"] + nav_groups, horizontal=True, label_visibility="collapsed"
        )

    def cat_in_group(cat, group):
        if group == "Tümü":
            return True
        path = cat.get("nav_path", [cat["title"]])
        return bool(path) and path[0] == group

    def gap_in_group(gap, group):
        if group == "Tümü":
            return True
        return gap.get("nav_group") == group

    cats_filtered = [c for c in cats if cat_in_group(c, selected_group)] if nav_groups else cats
    gaps = [g for g in gaps_all if gap_in_group(g, selected_group)] if nav_groups else gaps_all

    opp_count  = sum(1 for g in gaps if g["status"] == "opportunity")
    high_count = sum(1 for g in gaps if g["status"] == "high")

    st.markdown(f"""
    <div class="metric-grid">
        <div class="metric-card"><div class="label">Mevcut Kategori</div><div class="value">{len(cats_filtered)}</div><div class="sub">{len(noise_cats)} kampanya/test ayrıldı</div></div>
        <div class="metric-card"><div class="label">Toplam Ürün</div><div class="value">{len(products):,}</div><div class="sub">sitede taranan</div></div>
        <div class="metric-card"><div class="label">Analiz Edilen Keyword</div><div class="value">{len(kws):,}</div><div class="sub">organik keyword</div></div>
        <div class="metric-card good"><div class="label">Gerçek Fırsat</div><div class="value">{opp_count}</div><div class="sub">ürün var, sayfa yok</div></div>
        <div class="metric-card warn"><div class="label">Ürün Yok</div><div class="value">{high_count}</div><div class="sub">hacim var, ürün yok</div></div>
    </div>
    """, unsafe_allow_html=True)

    col_tree, col_bar = st.columns([3, 2])
    with col_tree:
        st.markdown('<div class="section-title">Kategori Haritası</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="legend">
            <span><span class="dot" style="background:#cdeedb"></span>Mevcut kategori</span>
            <span><span class="dot" style="background:#bfe8d2"></span>Fırsat (ürün var)</span>
            <span><span class="dot" style="background:#fad9b3"></span>Hacim var, ürün yok</span>
        </div>
        """, unsafe_allow_html=True)
        if cats_filtered or gaps:
            st.plotly_chart(build_treemap(cats_filtered, gaps), use_container_width=True, config={"displayModeBar": False})
        else:
            st.info("Bu grupta gösterilecek kategori verisi yok.")

    with col_bar:
        st.markdown('<div class="section-title">Gap Hacimleri</div>', unsafe_allow_html=True)
        bar = build_volume_bar(gaps)
        if bar:
            st.plotly_chart(bar, use_container_width=True, config={"displayModeBar": False})
        else:
            st.info("Bu grupta gap önerisi bulunamadı.")

    st.markdown(f'<div class="section-title">Kategori Önerileri <span class="count">{len(gaps)}</span></div>', unsafe_allow_html=True)
    if gaps:
        rows_html = "<div class='dtable'><div class='row head' style='grid-template-columns:2fr 1.3fr 1fr 1fr 1.4fr'>" \
                    "<span>Önerilen Kategori</span><span>Baz Kategori</span><span>Toplam Hacim</span>" \
                    "<span>Eşleşen Ürün</span><span>Durum</span></div>"
        for g in gaps[:60]:
            badge = f'<span class="badge {STATUS_BADGE_CLASS[g["status"]]}">{STATUS_LABEL[g["status"]]}</span>'
            rows_html += (
                f"<div class='row' style='grid-template-columns:2fr 1.3fr 1fr 1fr 1.4fr'>"
                f"<span class='cat-name'>+ {g['suggested_category']}</span>"
                f"<span class='mono'>{g['base_category']}</span>"
                f"<span class='mono'>{g['total_search_volume']:,}</span>"
                f"<span class='mono'>{g['matching_product_count']}</span>"
                f"<span>{badge}</span></div>"
                f"<div class='sample-kw'>{g['sample_keywords']}</div>"
            )
        rows_html += "</div>"
        st.markdown(rows_html, unsafe_allow_html=True)
        if len(gaps) > 60:
            st.caption(f"İlk 60 öneri gösteriliyor. Tam liste ({len(gaps)} satır) Excel raporunda.")
    else:
        st.info("Bu grupta hiç gap önerisi bulunamadı.")

    # ── En çok hacim getiren mevcut kategoriler ──
    top_cats_table = build_top_categories_table(cats_filtered, r.get("matched", []))
    st.markdown('<div class="section-title">En Çok Hacim Getiren Mevcut Kategoriler</div>', unsafe_allow_html=True)
    if top_cats_table:
        tc_html = "<div class='dtable'><div class='row head' style='grid-template-columns:2.5fr 1fr 1fr 1fr'>" \
                  "<span>Kategori</span><span>Ürün Sayısı</span><span>Toplam Hacim</span><span>Eşleşen Keyword</span></div>"
        for t in top_cats_table:
            tc_html += (
                f"<div class='row' style='grid-template-columns:2.5fr 1fr 1fr 1fr'>"
                f"<span class='cat-name'>{t['title']}</span>"
                f"<span class='mono'>{t['product_count']}</span>"
                f"<span class='mono'>{t['total_volume']:,}</span>"
                f"<span class='mono'>{t['keyword_count']}</span></div>"
            )
        tc_html += "</div>"
        st.markdown(tc_html, unsafe_allow_html=True)
    else:
        st.info("Bu grupta eşleşen keyword bulunamadı.")

    # ── Filtrelenen kampanya/test koleksiyonları — grid görünüm ──
    if noise_cats:
        st.markdown(f'<div class="section-title">Filtrelenen Kampanya / Test Koleksiyonları <span class="count">{len(noise_cats)}</span></div>', unsafe_allow_html=True)
        with st.expander("Listeyi göster"):
            chips = "".join(
                f"<span style='display:inline-block;background:#eceef3;color:#4a4e5c;font-size:0.78rem;"
                f"padding:4px 11px;border-radius:14px;margin:3px'>{c['title']}</span>"
                for c in noise_cats[:150]
            )
            st.markdown(f"<div>{chips}</div>", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:20px'>", unsafe_allow_html=True)
    excel_buf = generate_excel(gaps_all, cats, noise_cats, kws, dom)
    st.download_button(
        label="⬇ Excel Raporu İndir",
        data=excel_buf,
        file_name=f"seo_gap_{dom.replace('.','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown("</div>", unsafe_allow_html=True)
